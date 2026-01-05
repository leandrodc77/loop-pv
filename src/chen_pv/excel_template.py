"""
Leitura/escrita do template Excel (.xlsm) usado como base.

Notas importantes:
- `openpyxl` NÃO recalcula fórmulas. Para extrair resultados corretos de células com fórmula,
  você precisa abrir e salvar o arquivo no Excel/LibreOffice antes, OU calcular via `chen_pv.core`.
- Mantém macros (VBA) quando `keep_vba=True`.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Optional

import openpyxl


TEMPLATE_SHEET = "Planilha2"


@dataclass(frozen=True)
class TemplateVariable:
    label: str
    cell: str          # coordenada, ex.: "B2"
    unit: Optional[str]
    kind: str          # "input" ou "derived"
    formula: Optional[str]


def _load_ws(path: str, *, data_only: bool, keep_vba: bool = True):
    wb = openpyxl.load_workbook(path, data_only=data_only, keep_vba=keep_vba)
    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(f"Aba {TEMPLATE_SHEET!r} não encontrada em {path!r}. Encontradas: {wb.sheetnames}")
    return wb, wb[TEMPLATE_SHEET]


def extract_schema(path: str) -> List[TemplateVariable]:
    """
    Lê a tabela (colunas A-C) e devolve um "schema" das variáveis do template:
    - label (coluna A)
    - cell (coluna B)
    - unit (coluna C)
    - kind: input vs derived (derivado se coluna B começa com '=')
    - formula (se derivado)
    """
    wb, ws = _load_ws(path, data_only=False)
    out: List[TemplateVariable] = []

    # Tabela começa na linha 2 (linha 1 é cabeçalho)
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None or str(label).strip() == "":
            continue

        val = ws.cell(r, 2).value
        unit = ws.cell(r, 3).value

        formula = val if isinstance(val, str) and val.startswith("=") else None
        kind = "derived" if formula else "input"

        out.append(
            TemplateVariable(
                label=str(label).strip(),
                cell=f"B{r}",
                unit=None if unit is None else str(unit).strip(),
                kind=kind,
                formula=formula,
            )
        )

    return out


def extract_values(path: str) -> Dict[str, Optional[float]]:
    """
    Extrai (label -> valor) usando `data_only=True` (valores cacheados).
    """
    wb, ws = _load_ws(path, data_only=True)
    out: Dict[str, Optional[float]] = {}

    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None or str(label).strip() == "":
            continue
        val = ws.cell(r, 2).value
        out[str(label).strip()] = None if val is None else float(val)

    return out


def fill_template(
    *,
    template_path: str,
    output_path: str,
    inputs: Dict[str, float],
) -> None:
    """
    Preenche a coluna B do template para as variáveis de entrada e salva uma cópia.

    Chaves aceitas em `inputs` (canônicas):
    - PAS, PAD, VDF, VSF, PET, ET, Ees, V0

    Observação: no template a linha do PET é rotulada "PET (CIV)".
    """
    wb, ws = _load_ws(template_path, data_only=False)

    # Monta índice label->linha
    label_to_row: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        label_to_row[str(label).strip()] = r

    # Mapeia chaves canônicas para o label no Excel
    key_to_label = {
        "PAS": "PAS",
        "PAD": "PAD",
        "VDF": "VDF",
        "VSF": "VSF",
        "PET": "PET (CIV)",
        "ET": "ET",
        "Ees": "Ees",
        "V0": "V0",
    }

    for key, value in inputs.items():
        if key not in key_to_label:
            raise ValueError(f"Chave desconhecida em inputs: {key!r}. Aceitas: {sorted(key_to_label)}")
        label = key_to_label[key]
        if label not in label_to_row:
            raise ValueError(f"Label {label!r} não encontrado no template {template_path!r}.")
        r = label_to_row[label]
        ws.cell(r, 2).value = float(value)

    wb.save(output_path)
